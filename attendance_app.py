from __future__ import annotations

import tkinter as tk
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parent
EXCEL_FILE = APP_DIR / "attendance.xlsx"
STANDARD_HOURS = 8.0

HEADERS = [
    "Date",
    "Project Name",
    "Sub Contractor Name",
    "Sub Contractor ID",
    "Worker ID",
    "Worker Name",
    "Start Time",
    "End Time",
    "Break Time",
    "Work Hours",
    "Normal Hours",
    "OT Hours",
]


def parse_clock_time(value: str, field_name: str) -> datetime:
    cleaned = value.strip()
    for fmt in ("%H:%M", "%I:%M %p", "%I:%M%p"):
        try:
            parsed = datetime.strptime(cleaned, fmt)
            return datetime.combine(date.today(), parsed.time())
        except ValueError:
            pass
    raise ValueError(f"{field_name} must be like 08:00, 17:30, or 5:30 PM.")


def parse_break_hours(value: str) -> float:
    cleaned = value.strip()
    if not cleaned:
        return 0.0

    if ":" in cleaned:
        parts = cleaned.split(":")
        if len(parts) != 2:
            raise ValueError("Break time must be like 01:00 or 0.5.")
        hours, minutes = int(parts[0]), int(parts[1])
        if hours < 0 or minutes < 0 or minutes >= 60:
            raise ValueError("Break time minutes must be between 00 and 59.")
        return hours + minutes / 60

    hours = float(cleaned)
    if hours < 0:
        raise ValueError("Break time cannot be negative.")
    return hours


def calculate_hours(start_value: str, end_value: str, break_value: str) -> tuple[float, float, float]:
    start = parse_clock_time(start_value, "Start time")
    end = parse_clock_time(end_value, "End time")
    if end < start:
        end += timedelta(days=1)

    break_hours = parse_break_hours(break_value)
    work_hours = round(((end - start).total_seconds() / 3600) - break_hours, 2)
    if work_hours < 0:
        raise ValueError("Break time cannot be longer than total shift time.")

    normal_hours = round(min(work_hours, STANDARD_HOURS), 2)
    ot_hours = round(max(0.0, work_hours - STANDARD_HOURS), 2)
    return work_hours, normal_hours, ot_hours


def ensure_workbook() -> None:
    if EXCEL_FILE.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for index, header in enumerate(HEADERS, start=1):
        width = max(len(header) + 3, 14)
        sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(EXCEL_FILE)


def append_attendance(row: list[str | float]) -> None:
    ensure_workbook()
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook["Attendance"]
    sheet.append(row)
    workbook.save(EXCEL_FILE)


def load_existing_rows() -> list[list[str]]:
    ensure_workbook()
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook["Attendance"]
    rows: list[list[str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(value is not None for value in row):
            rows.append(["" if value is None else str(value) for value in row])
    return rows


class AttendanceApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Staff Attendance Entry")
        self.geometry("1180x720")
        self.minsize(980, 620)

        self.fields: dict[str, tk.StringVar] = {
            "Date": tk.StringVar(value=date.today().isoformat()),
            "Project Name": tk.StringVar(),
            "Sub Contractor Name": tk.StringVar(),
            "Sub Contractor ID": tk.StringVar(),
            "Worker ID": tk.StringVar(),
            "Worker Name": tk.StringVar(),
            "Start Time": tk.StringVar(value="08:00"),
            "End Time": tk.StringVar(value="17:00"),
            "Break Time": tk.StringVar(value="1:00"),
        }

        self.result_text = tk.StringVar(value="Work Hours: 0 | Normal: 0 | OT: 0")

        self.configure(padx=18, pady=18)
        self.create_form()
        self.create_table()
        self.refresh_table()

    def create_form(self) -> None:
        title = ttk.Label(self, text="Staff Attendance Entry", font=("Segoe UI", 18, "bold"))
        title.grid(row=0, column=0, sticky="w", columnspan=4, pady=(0, 14))

        form = ttk.Frame(self)
        form.grid(row=1, column=0, sticky="ew")
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        labels = list(self.fields.keys())
        for index, label in enumerate(labels):
            row = index // 2
            column = (index % 2) * 2
            ttk.Label(form, text=label).grid(row=row, column=column, sticky="w", padx=(0, 8), pady=6)
            ttk.Entry(form, textvariable=self.fields[label]).grid(
                row=row,
                column=column + 1,
                sticky="ew",
                padx=(0, 22),
                pady=6,
            )

        buttons = ttk.Frame(self)
        buttons.grid(row=2, column=0, sticky="ew", pady=(14, 10))

        ttk.Button(buttons, text="Calculate", command=self.calculate_preview).pack(side="left", padx=(0, 8))
        ttk.Button(buttons, text="Save to Excel", command=self.save_record).pack(side="left", padx=(0, 8))
        ttk.Button(buttons, text="Clear Form", command=self.clear_form).pack(side="left", padx=(0, 8))
        ttk.Label(buttons, textvariable=self.result_text, font=("Segoe UI", 10, "bold")).pack(
            side="left",
            padx=(18, 0),
        )

        note = ttk.Label(
            self,
            text="Time format examples: 08:00, 17:30, 5:30 PM. Break time examples: 1:00 or 0.5.",
        )
        note.grid(row=3, column=0, sticky="w", pady=(0, 10))

    def create_table(self) -> None:
        table_frame = ttk.Frame(self)
        table_frame.grid(row=4, column=0, sticky="nsew")
        self.rowconfigure(4, weight=1)
        self.columnconfigure(0, weight=1)

        self.table = ttk.Treeview(table_frame, columns=HEADERS, show="headings", height=14)
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.table.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.table.xview)
        self.table.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        for header in HEADERS:
            self.table.heading(header, text=header)
            self.table.column(header, width=120, minwidth=90, anchor="center")

        self.table.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

    def calculate_preview(self) -> tuple[float, float, float] | None:
        try:
            work_hours, normal_hours, ot_hours = calculate_hours(
                self.fields["Start Time"].get(),
                self.fields["End Time"].get(),
                self.fields["Break Time"].get(),
            )
        except ValueError as error:
            messagebox.showerror("Check time entry", str(error))
            return None

        self.result_text.set(
            f"Work Hours: {work_hours:.2f} | Normal: {normal_hours:.2f} | OT: {ot_hours:.2f}"
        )
        return work_hours, normal_hours, ot_hours

    def save_record(self) -> None:
        required_fields = [
            "Date",
            "Project Name",
            "Sub Contractor Name",
            "Sub Contractor ID",
            "Worker ID",
            "Worker Name",
            "Start Time",
            "End Time",
        ]
        missing = [field for field in required_fields if not self.fields[field].get().strip()]
        if missing:
            messagebox.showerror("Missing details", "Please enter: " + ", ".join(missing))
            return

        totals = self.calculate_preview()
        if totals is None:
            return

        work_hours, normal_hours, ot_hours = totals
        row: list[str | float] = [
            self.fields["Date"].get().strip(),
            self.fields["Project Name"].get().strip(),
            self.fields["Sub Contractor Name"].get().strip(),
            self.fields["Sub Contractor ID"].get().strip(),
            self.fields["Worker ID"].get().strip(),
            self.fields["Worker Name"].get().strip(),
            self.fields["Start Time"].get().strip(),
            self.fields["End Time"].get().strip(),
            self.fields["Break Time"].get().strip(),
            work_hours,
            normal_hours,
            ot_hours,
        ]

        append_attendance(row)
        self.refresh_table()
        messagebox.showinfo("Saved", f"Attendance saved to:\n{EXCEL_FILE}")

    def clear_form(self) -> None:
        keep_date = date.today().isoformat()
        for field, variable in self.fields.items():
            variable.set("")
        self.fields["Date"].set(keep_date)
        self.fields["Start Time"].set("08:00")
        self.fields["End Time"].set("17:00")
        self.fields["Break Time"].set("1:00")
        self.result_text.set("Work Hours: 0 | Normal: 0 | OT: 0")

    def refresh_table(self) -> None:
        self.table.delete(*self.table.get_children())
        for row in load_existing_rows():
            self.table.insert("", "end", values=row)


if __name__ == "__main__":
    ensure_workbook()
    app = AttendanceApp()
    app.mainloop()
